import csv
from openpyxl import load_workbook

def xlsx_to_csv(xlsx_file, csv_file):
    """
    Converts an XLSX file to a CSV file.

    Args:
        xlsx_file (str): Path to the XLSX file.
        csv_file (str): Path to the output CSV file.
    """
    try:
        workbook = load_workbook(xlsx_file, read_only=True)  # Optimize for large files
        sheet = workbook.active  # Get the active sheet

        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for row in sheet.iter_rows():
                csv_row = [cell.value for cell in row]
                writer.writerow(csv_row)

        print(f"Successfully converted '{xlsx_file}' to '{csv_file}'")

    except FileNotFoundError:
        print(f"Error: File '{xlsx_file}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
xlsx_to_csv('psgc_data.xlsx', 'psgc_data.csv')